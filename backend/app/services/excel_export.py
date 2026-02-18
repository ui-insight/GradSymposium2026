"""Excel export service for results."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.results import ProjectResult


def export_results_to_excel(
    poster_results: list[ProjectResult],
    art_results: list[ProjectResult],
) -> io.BytesIO:
    """Generate an Excel workbook with ranked results."""
    wb = Workbook()

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    columns = [
        "Rank",
        "Project #",
        "Title",
        "Presenter",
        "Department",
        "# Judges",
        "Total Score",
        "Avg Score",
    ]

    def _write_sheet(ws, title: str, results: list[ProjectResult]):
        ws.title = title

        # Header row
        for col, header in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=r.Rank)
            ws.cell(row=row_idx, column=2, value=r.Project_Number)
            ws.cell(row=row_idx, column=3, value=r.Project_Title)
            ws.cell(
                row=row_idx,
                column=4,
                value=f"{r.Presenter_First_Name} {r.Presenter_Last_Name}",
            )
            ws.cell(row=row_idx, column=5, value=r.Department or "")
            ws.cell(row=row_idx, column=6, value=r.Judge_Count)
            ws.cell(row=row_idx, column=7, value=r.Total_Score)
            ws.cell(row=row_idx, column=8, value=r.Average_Score)

        # Auto-width columns
        for col in range(1, len(columns) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = min(max_len + 4, 50)

    # Poster sheet
    ws_poster = wb.active
    _write_sheet(ws_poster, "Poster Rankings", poster_results)

    # Art sheet
    ws_art = wb.create_sheet()
    _write_sheet(ws_art, "Art Rankings", art_results)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
